from tkinter import*
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib



root=Tk()
root.title("Students Data Entry")
root.geometry('650x590')
root.resizable(False,False)
root.config(bg="#192a56")


file=pathlib.Path('Backend_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Admission No"
    sheet['B1']="Name"
    sheet['C1']="Father's name"
    sheet['D1']="Class"
    sheet['E1']="Age"
    sheet['F1']="Gender"
    sheet['G1']="Contact"
    sheet['H1']="Address"
    
    file.save('Backend_data.xlsx')


def submit():
    print("Submit function called")


    admission = admissionValue.get()
    name = nameValue.get()
    fathername = fathernameValue.get()
    Class = ClassValue.get()
    age = ageValue.get()
    contact = contactValue.get()
    gender = gender_combobox.get()
    address = addressEntry.get(1.0, END)


    print("Admission No :", admission)
    print("Name         :", name)
    print("Father's name:", fathername)
    print("Class        :", Class)
    print("Age          :", age)
    print("Gender       :", gender)
    print("Contact      :", contact)
    print("Address      :", address)

    file=openpyxl.load_workbook('Backend_data.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=admission)
    sheet.cell(column=2,row=sheet.max_row,value=name)
    sheet.cell(column=3,row=sheet.max_row,value=fathername)
    sheet.cell(column=4,row=sheet.max_row,value=Class)
    sheet.cell(column=5,row=sheet.max_row,value=age)
    sheet.cell(column=6,row=sheet.max_row,value=gender)
    sheet.cell(column=7,row=sheet.max_row,value=contact)
    sheet.cell(column=8,row=sheet.max_row,value=address)
    
    file.save(r'Backend_data.xlsx')
    
    
    messagebox.showinfo('info','detail added!')
    
    
    admissionValue.set('')
    nameValue.set('')
    contactValue.set('')
    fathernameValue.set('')
    ClassValue.set('')
    ageValue.set('')
    addressEntry.delete(1.0,END)
    gender_combobox.set('')
    


def clear():
    admissionValue.set('')
    nameValue.set('')
    contactValue.set('')
    fathernameValue.set('')
    ClassValue.set('')
    ageValue.set('')
    addressEntry.delete(1.0,END)
    gender_combobox.set('')

icon_image=PhotoImage(file="logo.png.png")
root.iconphoto(False,icon_image)


Label(root,text="ENTRY FORM ",font='arial 25',bg="#718093",fg="#ecf0f1").place(x=230,y=20)


Label(root,text="Admission No :",font='arial 15',bg="#192a56",fg="#ecf0f1").place(x=50,y=100)
Label(root,text="Name :",font='arial 15',bg="#192a56",fg="#ecf0f1").place(x=50,y=150)
Label(root,text="Father's name :",font='arial 15',bg="#192a56",fg="#ecf0f1").place(x=50,y=200)
Label(root,text="Class :",font='arial 15',bg="#192a56",fg="#ecf0f1").place(x=50,y=250)
Label(root,text="Age :",font='arial 15',bg="#192a56",fg="#ecf0f1").place(x=410,y=250)
Label(root,text="Gender :",font='arial 15',bg="#192a56",fg="#ecf0f1").place(x=50,y=300)
Label(root,text="Contact :",font='arial 15',bg="#192a56",fg="#ecf0f1").place(x=50,y=350)
Label(root,text="Address :",font='arial 15',bg="#192a56",fg="#ecf0f1").place(x=50,y=400)

admissionValue=StringVar()
nameValue = StringVar()
fathernameValue = StringVar()
contactValue = StringVar()
addressValue = StringVar()
ClassValue=StringVar()
ageValue=StringVar()

admissionEntry = Entry(root, textvariable=admissionValue, width=7, bd=3, font='arial 15')
admissionEntry.place(x=200, y=100)

nameEntry = Entry(root, textvariable=nameValue, width=40, bd=2, font=23)
nameEntry.place(x=200, y=150)

fathernameEntry = Entry(root, textvariable=fathernameValue, width=40, bd=2, font=20)
fathernameEntry.place(x=200, y=200)

classEntry = Entry(root, textvariable=ClassValue, width=5 ,bd=3, font='arial 15')
classEntry.place(x=200, y=250)

ageEntry = Entry(root, textvariable=ageValue, width=3, bd=3, font='arial 15')
ageEntry.place(x=470, y=250)

gender_combobox = Combobox(root, values=['Male', 'Female'], font='arial 14', state='readonly', width=14)
gender_combobox.place(x=200, y=300)

contactEntry = Entry(root, textvariable=contactValue, width=17, bd=2, font='arial 15')
contactEntry.place(x=200, y=350)

addressEntry = Text(root, width=40, height=4, bd=4,font='arial 13')
addressEntry.place(x=200, y=400)


Button(root,text="Exit",bg="#ED4C67",activebackground="#B53471",fg="#ecf0f1",width=15,bd=3,height=2,command=lambda:root.destroy()).place(x=175,y=520)
Button(root,text="Clear",bg="#22a6b3",activebackground="#006266",fg="#ecf0f1",width=15,bd=3,height=2,command=clear).place(x=50,y=520)
Button(root,text="Submit",bg="#44bd32",activebackground="#4cd137",fg="#ecf0f1",width=20,height=2,bd=3,command=submit).place(x=460,y=520)













root.mainloop()